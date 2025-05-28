
from openpyxl import load_workbook

from database.connection import Base, engine, session
from database.models.Material import MaterialModel
from database.models.Order import OrderModel
from database.models.Partner import PartnerModel
from database.models.Product import ProductModel
from database.models.ProductType import ProductTypeModel

Base.metadata.drop_all(engine)
Base.metadata.create_all(engine)

product_type_file = "excel/Product_type_import.xlsx"
products_file = "excel/Products_import.xlsx"
partners_file = "excel/Partners_import.xlsx"
sales_history_file = "excel/Partner_products_import.xlsx"
materials_file = "excel/Material_type_import.xlsx"


def import_product_types(session):
    wb = load_workbook(product_type_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, coefficient = row
        if not type or not coefficient:
            continue
        session.add(ProductTypeModel(type=type, coefficient=coefficient))
    session.commit()
    print("Типы продукции добавлены.")


def import_partners(session):
    wb = load_workbook(partners_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, company_name, boss_name, mail, phone_number, address, inn, rank = row
        session.add(PartnerModel(
            type=type,
            company_name=company_name,
            address=address,
            inn=inn,
            boss_name=boss_name,
            phone_number=phone_number,
            mail=mail,
            rank=rank
        ))
    session.commit()
    print("Партнёры добавлены.")


def import_materials(session):
    wb = load_workbook(materials_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, percentage_of_defective_material = row
        session.add(MaterialModel(
            type=type,
            percentage_of_defective_material=percentage_of_defective_material
        ))
    session.commit()
    print("Материалы добавлены.")


def import_sales_history(session):
    wb = load_workbook(sales_history_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, partner_name, quantity_of_products, date_of_create = row
        product = session.query(ProductModel).filter_by(name=product_name).first()
        partner = session.query(PartnerModel).filter_by(company_name=partner_name).first()
        session.add(OrderModel(
            fk_product_id=product.id,
            fk_company_id=partner.id,
            quantity_of_products=quantity_of_products,
            date_of_create=date_of_create
        ))
    session.commit()
    print("История продаж добавлена.")


import_product_types(session)


def import_products(session):
    wb = load_workbook(products_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fk_type, name, article, min_cost = row
        product_type = session.query(ProductTypeModel).filter_by(type=fk_type).first()
        session.add(ProductModel(
            article=article,
            name=name,
            min_cost=min_cost,
            fk_type=product_type.id
        ))
    session.commit()
    print("Продукты добавлены.")

import_products(session)
import_partners(session)
import_sales_history(session)
import_materials(session)
